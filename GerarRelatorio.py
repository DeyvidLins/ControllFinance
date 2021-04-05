from openpyxl import Workbook # pip install openpyxl
from openpyxl.styles import Font, Color, colors
import ConexaoBD




class GerarPlanilha():
    def gerar_pla(self,data, dir):
        wb = Workbook()
        ws1 = wb.active # work sheet
        ws1.title = "Pyxl"

        listDesc=[]
        listValor=[]
        listData=[]
        listPag=[]
        listValorPag=[]
        listDev=[]
        listStatus=[]

        ws1["A1"]='Descrição'
        ws1["B1"]='Valor'
        ws1["C1"]='Data de Vencimento'
        ws1["D1"]='Data de Pagamento'
        ws1["E1"]='Valor que foi pago'
        ws1["F1"]='Devendo'
        ws1["G1"]='Status'

        conection= ConexaoBD.cur

        #Laços de interação para pegar as informações do banco e transformar em uma planilha
        for i in conection.execute(f"select desc from finance WHERE dataVenc LIKE '%{data}'").fetchall():
            listDesc.append(i[0])
            ws1.cell(column=1, row=len(listDesc) + 1 , value=i[0])


        for i in conection.execute(f"SELECT valor  FROM finance WHERE dataVenc LIKE '%{data}' ").fetchall():
            listValor.append(i[0])
            ws1.cell(column=2, row=len(listValor) + 1, value=i[0])


        for i in conection.execute(f"SELECT dataVenc FROM finance WHERE dataVenc LIKE '%{data}' ").fetchall():
            listData.append(i[0])
            ws1.cell(column=3, row=len(listData) + 1, value=i[0])

        for i in conection.execute(f"SELECT dataPag FROM finance WHERE dataVenc LIKE '%{data}' ").fetchall():
            listPag.append(i[0])
            ws1.cell(column=4, row=len(listPag) + 1, value=i[0])

        for i in conection.execute(f"SELECT valorPag FROM finance WHERE dataVenc LIKE '%{data}' ").fetchall():
            listValorPag.append(i[0])
            ws1.cell(column=5, row=len(listValorPag) + 1, value=i[0])

        for i in conection.execute(f"select devendo from finance WHERE dataVenc LIKE '%{data}' ").fetchall():
            listDev.append(i[0])
            ws1.cell(column=6, row=len(listDev) + 1, value=i[0])

        for i in conection.execute(f"select status from finance WHERE dataVenc LIKE '%{data}' ").fetchall():
            listStatus.append(i[0])
            ws1.cell(column=7, row=len(listStatus) + 1, value=i[0])


        #Inserindo dados no FireBase
        cont = 0


        ws1["H13"]='TOTAL DE DÍVIDA'
        ws1["I13"]='=SUM(F2:F50)'


        ft_a = Font(name='Times New Roman',color=colors.BLACK, bold=True, size=12)

        a1 = ws1['A1']
        b1 = ws1['B1']
        c1 = ws1['C1']
        d1 = ws1['D1']
        e1 = ws1['E1']
        f1 = ws1['F1']
        g1 = ws1['G1']
        h13 = ws1['H13']
        a1.font = ft_a
        b1.font = ft_a
        c1.font = ft_a
        d1.font = ft_a
        e1.font = ft_a
        f1.font = ft_a
        g1.font = ft_a
        h13.font = ft_a

        wb.save(f'{dir}')

