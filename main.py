from openpyxl import Workbook # pip install openpyxl
from openpyxl.styles import Font, Color, colors
import  ConexaoBD
import pandas as pd

wb = Workbook()
ws1 = wb.active # work sheet
ws1.title = "Pyxl"

#Realiza a leitura  do arquvio excel através da biblioteca Pandas
#x = pd.read_excel(r"C:\Users\Deyvid\Desktop\Repo-GitHub\pythonExcel\ControleFinanceiro.xlsx")
#print(x)

listDesc=[]
listValor=[]
listData=[]
listPag=[]
listValorPag=[]
listDev=[]
listStatus=[]

ws1["A1"]='Descrição'
ws1["B1"]='Valor'
ws1["C1"]='Data de Vencimento'
ws1["D1"]='Data de Pagamento'
ws1["E1"]='Valor que foi pago'
ws1["F1"]='Devendo'
ws1["G1"]='Status'

conection= ConexaoBD.cur
conectionFireBase = ConexaoBD.bd



for i in conection.execute("select desc from financa; ").fetchall():
    listDesc.append(i[0])
    ws1.cell(column=1, row=len(listDesc) + 1 , value=i[0])


for i in conection.execute("select valor from financa; ").fetchall():
    listValor.append(i[0])
    ws1.cell(column=2, row=len(listValor) + 1, value=i[0])


for i in conection.execute("select dataVenc from financa; ").fetchall():
    listData.append(i[0])
    ws1.cell(column=3, row=len(listData) + 1, value=i[0])

for i in conection.execute("select dataPag from financa; ").fetchall():
    listPag.append(i[0])
    ws1.cell(column=4, row=len(listPag) + 1, value=i[0])

for i in conection.execute("select valorPag from financa; ").fetchall():
    listValorPag.append(i[0])
    ws1.cell(column=5, row=len(listValorPag) + 1, value=i[0])

for i in conection.execute("select devendo from financa; ").fetchall():
    listDev.append(i[0])
    ws1.cell(column=6, row=len(listDev) + 1, value=i[0])

for i in conection.execute("select status from financa; ").fetchall():
    listStatus.append(i[0])
    ws1.cell(column=7, row=len(listStatus) + 1, value=i[0])

cont = 0
for i in conection.execute("select * from financa; ").fetchall():
    conectionFireBase.child("Finance").push({"Descrição": f"{listDesc[cont]}","Valor": f"{listValor[cont]}" , "Data de Vencimento": f"{listData[cont]}",
                       "Data de Pagamento": f"{listPag[cont]}", "Valor que foi pago": f"{listValorPag[cont]}", "Devendo": f"{listDev[cont]}",
                        "Status": f"{listStatus[cont]}"})
    cont += 1


ws1["H13"]='TOTAL DE DÍVIDA'
ws1["I13"]='=SUM(F2:F50)'


ft_a = Font(name='Times New Roman',color=colors.BLACK, bold=True, size=12)

a1 = ws1['A1']
b1 = ws1['B1']
c1 = ws1['C1']
d1 = ws1['D1']
e1 = ws1['E1']
f1 = ws1['F1']
g1 = ws1['G1']
h13 = ws1['H13']
a1.font = ft_a
b1.font = ft_a
c1.font = ft_a
d1.font = ft_a
e1.font = ft_a
f1.font = ft_a
g1.font = ft_a
h13.font = ft_a

wb.save('ControleFinanceiro.xlsx')
